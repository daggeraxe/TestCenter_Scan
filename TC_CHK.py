import csv
import time
import os
import openpyxl
from openpyxl import load_workbook
import TC_GET as tget
import TC_SCAN as scan

timestamp = time.strftime('%Y-%m-%d-%H-%M-%S')
oldwb = time.strftime('%Y%m%d')
csvpath = 'D:\\TestCenter\\equipment\\equipment' + timestamp + '.csv'
# csvpath = 'D:\\TestCenter\\equipment\\equipment2020-11-04_14-39-14.csv'
excel_file = 'D:\\TestCenter\\STC_SHA6_gzhao.xlsx'
heard_name = ['Chassis', 'TestModule', 'PortGroup', 'Diagnostics']
heards = []
zone = [[], [], [], []]
chassis_info = dict()
module_info = dict()
datas = dict()
chassis_dic = dict()
module_dic = dict()
modulename_dic = {
    'EDM-1001B': '4Port 10/100/100 DUAL MEDIA',
    'EDM-1003B': '12PORT 10/100/1000 DUAL MEDIA',
    'EDM-2003B': '12PORT 10/100/100 DUAL MEDIA',
    'MSA-1001B': '2 PORT 10GbE MULTI-MSA HOST',
    'CV-10G-S2': 'hyperMETRICS CV 2 PORT 10G SFP+',
    'CV-10G-S8': 'hyperMETRICS CV 8 PORT 10G SFP+',
    'DX-10G-S32': 'hyperMETRICS DX 32 PORT 10G SFP+',
    'FX2-1G-S16': 'FX2 16-PORT 1GBE SFP',
    'FX2-10G-S12': '12 PORT 1G/10G DUAL RATE',
    'FX2-10G-S16': '16 PORT 1G/10G  DUAL RATE',
    'FX2-10G-Q2': 'FX 40GE 2 ports',
    'FX-100G-P2': 'FX 100GBE ONLY CFP2-2-PORTS',
    'FX3-100GTL-T2': 'FX3 2-PORT 100/40/10GBE QSFP28',
    'DX2-100G-P4': 'DX2 4-PORT 100/40GBE CFP2',
    'DX2-100GO-P4': 'DX2 4-PORT 100GBE CFP2',
    'DX2-100GO-T8': 'DX2 8-PORT100GBE QSFP28',
    'ACC-6084A': '100G/40G CFP2 (AFBR-8420Z)'
}

data_key = ('Type', 'Name', 'Controller S/N', 'Program/Team', 'User',
            'Location/IP', 'Checked IP', 'Comments')
chassisoffline_ip = list()
x = -1  # for CSV file splite zone nomber, start with -1
stc_maxrow = 40  # STC sheet max row
mod_maxrow = 77  # module sheet max row

# start timer
print('[+] Start Testcenter Checking')
t1 = time.time()

# Scan specified networks for find Testcenters
networks = [
    '172.29.93.0/24', '172.29.97.0/24', '172.29.201.0/24', '172.29.52.0/24',
    '172.29.92.0/24', '172.29.49.0/24', '172.29.96.0/24', '172.29.99.0/24',
    '172.29.157.0/24', '10.13.16.0/24', '10.13.15.0/24'
]
# networks = ['172.29.0.0/16', '10.13.0.0/16']
print('[+] Start Scan ...', networks)
for ipd in networks:
    scan.scan_network(ipd)

print('[+] Start get SN form file')
# read txt and get sn
raw_iplist = scan.get_last_line('TestCenter_IP_SCAN.txt')
szChassisIpList = []
for ipinlist in raw_iplist.split():
    szChassisIpList.append(str(ipinlist, 'utf-8'))
os.remove('TestCenter_IP_SCAN.txt')
# Write SN/PN to CSV
tget.get_sn_to_csv(szChassisIpList, csvpath)

# open excel file by openpyxl
wb = load_workbook(filename=excel_file)
try:
    ws = wb["database"]
    ws.title = oldwb
    # wb.remove(ws)
except Exception as ERR:
    print('ERROR -- ', ERR)
sheet_db = wb.create_sheet(title="database", index = 4)

# open csv file
with open(csvpath, 'r', newline='') as file:
    csvreader = csv.reader(file)
    # read csv file
    for row in csvreader:
        # print('read line: ', row, len(row))
        if '#' in row[0] and csvreader.line_num > 5:
            # skip first 3 lines and add equipment type to heards
            nextline = next(csvreader)
            # print(nextline)
            heards.append(nextline)
            x += 1
        else:
            # add contect to data named zone[0 .. 3], zone[0] for Chassis information
            zone[x].append(row)
for a, b in zip(heard_name, heards):
    locals()[a] = b

for aline in zone[0]:
    for key, value in zip(Chassis, aline):
        chassis_info.setdefault(key, []).append(value)

for aline in zone[1]:
    for key, value in zip(TestModule, aline):
        module_info.setdefault(key, []).append(value)

# Update data by openpyxl
isfirstline = 0
for ip, sn in zip(chassis_info['Chassis IP or Hostname'],
                  chassis_info['Serial Number']):
    if isfirstline == 0:
        sheet_db.append(['# Chassis'])
        sheet_db.append(['SN', 'IP'])
    sheet_db.append([sn, ip])
    chassis_dic.update({sn: ip})
    isfirstline += 1
isfirstline = 0
for ip, sn in zip(module_info['Chassis IP or Hostname'],
                  module_info['Serial Number']):
    if isfirstline == 0:
        sheet_db.append(['# Test Module'])
    sheet_db.append([sn, ip])
    module_dic.update({sn: ip})
    isfirstline += 1

ws_stc = wb["STC"]
chassis_sn = chassis_info['Serial Number']
sn_columns = list(
    ws_stc.iter_cols(min_col=7, max_col=7, min_row=2, max_row=stc_maxrow))[0]
st_columns = list(
    ws_stc.iter_cols(min_col=13, max_col=13, min_row=2, max_row=stc_maxrow))[0]
ip_columns = list(
    ws_stc.iter_cols(min_col=14, max_col=14, min_row=2, max_row=stc_maxrow))[0]
us_columns = list(
    ws_stc.iter_cols(min_col=12, max_col=12, min_row=2, max_row=stc_maxrow))[0]
ct_columns = list(
    ws_stc.iter_cols(min_col=1, max_col=1, min_row=2, max_row=stc_maxrow))[0]
pt_columns = list(
    ws_stc.iter_cols(min_col=11, max_col=11, min_row=2, max_row=stc_maxrow))[0]

print('[+] Checking Chassis SN')
for sn_column, st_column in zip(sn_columns, st_columns):
    current_row = sn_column.row - 2
    # print('checking STC row:', current_row)
    checkedip = chassis_dic.get(sn_column.value)
    if sn_column.value not in chassis_sn and st_column.value == 'in use':
        print('  [-] NOT Found', sn_column.value, st_column.value, 'row: ',
              current_row)
        data = (ct_columns[current_row].value, 'Chassis',
                sn_columns[current_row].value, pt_columns[current_row].value,
                us_columns[current_row].value, ip_columns[current_row].value,
                'NA', 'Connection Failed')
        # print('data:  ', data)
        chassisoffline_ip.append(ip_columns[current_row].value)
        for key, value in zip(data_key, data):
            datas.setdefault(key, []).append(value)
    elif sn_column.value in chassis_sn and checkedip != ip_columns[
            current_row].value:
        print('  [-] Found, but IP Changed', sn_column.value)
        data = (ct_columns[current_row].value, 'Chassis',
                sn_columns[current_row].value, pt_columns[current_row].value,
                us_columns[current_row].value, ip_columns[current_row].value,
                checkedip, 'IP Changed')
        # print('data:  ', data)
        for key, value in zip(data_key, data):
            datas.setdefault(key, []).append(value)
print('  [-] Chassis in list was not Found', chassisoffline_ip)
print('[+] Finished Check Chassis SN')

ws_mod = wb["STC Module"]
module_sn = module_info['Serial Number']
msn_columns = list(
    ws_mod.iter_cols(min_col=5, max_col=5, min_row=2, max_row=mod_maxrow))[0]
mst_columns = list(
    ws_mod.iter_cols(min_col=7, max_col=7, min_row=2, max_row=mod_maxrow))[0]
mip_columns = list(
    ws_mod.iter_cols(min_col=10, max_col=10, min_row=2, max_row=mod_maxrow))[0]
mus_columns = list(
    ws_mod.iter_cols(min_col=9, max_col=9, min_row=2, max_row=mod_maxrow))[0]
mmt_columns = list(
    ws_mod.iter_cols(min_col=1, max_col=1, min_row=2, max_row=mod_maxrow))[0]
mpt_columns = list(
    ws_mod.iter_cols(min_col=8, max_col=8, min_row=2, max_row=mod_maxrow))[0]

print('[+] Checking Module SN')
for sn_column, st_column, ip_column in zip(msn_columns, mst_columns,
                                           mip_columns):
    current_row = sn_column.row - 2
    checkedip = module_dic.get(sn_column.value)
    modulename = modulename_dic.get(mmt_columns[current_row].value)
    # print('  [-] Checking STC Module row', current_row,
    #       'get SN', sn_column.value, 'in Chassis', checkedip, ' under ', st_column.value)
    if ip_column.value in chassisoffline_ip and checkedip == None:
        print('  [-] NOT Found', sn_column.value, st_column.value, 'in row: ',
              current_row, 'Due to Chassis Not Found')
        data = (mmt_columns[current_row].value, modulename,
                msn_columns[current_row].value, mpt_columns[current_row].value,
                mus_columns[current_row].value, mip_columns[current_row].value,
                'NA', 'Chassis Offline')
        # print('data:  ', data)
        for key, value in zip(data_key, data):
            datas.setdefault(key, []).append(value)
    elif sn_column.value not in module_sn and st_column.value == 'in use':
        print('  [-] NOT Found', sn_column.value, st_column.value, 'in row: ',
              current_row)
        data = (mmt_columns[current_row].value, modulename,
                msn_columns[current_row].value, mpt_columns[current_row].value,
                mus_columns[current_row].value, mip_columns[current_row].value,
                'NA', 'Not detected')
        # print('data:  ', data)
        for key, value in zip(data_key, data):
            datas.setdefault(key, []).append(value)
    elif sn_column.value in module_sn and checkedip != mip_columns[
            current_row].value:
        print('  [-] Found, but IP Changed', sn_column.value)
        data = (mmt_columns[current_row].value, modulename,
                msn_columns[current_row].value, mpt_columns[current_row].value,
                mus_columns[current_row].value, mip_columns[current_row].value,
                checkedip, 'Chassis Changed')
        # print('data:  ', data)
        for key, value in zip(data_key, data):
            datas.setdefault(key, []).append(value)

cellx = 4
for column in list(datas.items()):
    celly = 1
    sheet_db.cell(column=cellx, row=celly, value=column[0])
    for row in column[1]:
        celly += 1
        sheet_db.cell(column=cellx, row=celly, value=row)
        # print(cellx, celly, row)
    cellx += 1

wb.save(filename=excel_file)
wb.close()

t2 = time.time()
print("[+] 本次扫描共花费 %s 秒" % (t2 - t1))
