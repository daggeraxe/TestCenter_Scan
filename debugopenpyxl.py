# debug openpyxl
import time
from openpyxl import load_workbook
excel_file = 'D:\\TestCenter\\STC_SHA6_gzhao.xlsx'
oldwb = time.strftime('%Y%m%d')
# open excel file by openpyxl
wb = load_workbook(filename=excel_file)
try:
    ws = wb["database"]
    ws.title = oldwb
    # wb.remove(ws)
except Exception as ERR:
    print('ERROR -- ', ERR)
sheet_db = wb.create_sheet(title="database", index = 4)
sheet_db['A1'] = 'You ARE WELCOME'
wb.save(filename=excel_file)
wb.close()